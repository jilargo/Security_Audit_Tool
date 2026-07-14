
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from pathlib import Path
from datetime import datetime
import pandas as pd
from collections import Counter   # Used for counting top processes


class SecurityAudit:
    """
    Core class responsible for generating a professional Security Audit Report.
    
    This class collects data from the GUI form and various system collectors,
    then compiles everything into a well-formatted, visually appealing Excel report
    with a dashboard, risk assessment, and charts.
    """

    @staticmethod
    def get_form_data(window):
        """
        Extract all the information entered by the user in the MainWindow GUI.
        
        Args:
            window: The MainWindow instance containing the input fields
            
        Returns:
            dict: Employee and audit context data
        """
        data = {
            "first_name": window.firstname_input.text().strip(),
            "last_name": window.lastname_input.text().strip(),
            "position": window.position_input.text().strip(),
            "company_id": window.company_id_input.text().strip(),
            "company_email": window.company_email_input.text().strip(),
            "date_hired": window.date_hired_input.date().toString("yyyy-MM-dd"),
            "Department": window.department_input.currentText(),
        }
        return data

    @staticmethod
    def _style_worksheet(ws, is_dashboard=False):
        """
        Apply professional styling to an Excel worksheet.
        
        This includes:
        - Colored headers
        - Borders
        - Auto-adjusting column widths
        - Freeze panes for easy scrolling
        - Text wrapping for readability
        
        Args:
            ws: openpyxl worksheet object
            is_dashboard: Whether this is the special Dashboard sheet (different styling)
        """
        if not ws or ws.max_row == 0:
            return
            
        # Professional dark blue header style
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Style the header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        if not is_dashboard:
            # Auto-adjust column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[column_letter].width = max_length + 2
            
            # Freeze the header row so it stays visible when scrolling
            ws.freeze_panes = "A2"
            
            # Style data rows
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.border = border

    @staticmethod
    def generate_audit_report(window, output_folder="reports", logo_path=None):
        """
        Main method that generates the complete security audit Excel report.
        
        Process flow:
        1. Collect employee data from GUI
        2. Gather system security information from various collectors
        3. Assess risk level based on findings
        4. Create Excel file with multiple sheets using pandas
        5. Apply advanced formatting, charts, and logo using openpyxl
        6. Save and return the file path
        
        Args:
            window: MainWindow instance
            output_folder: Folder where the report will be saved
            logo_path: Optional path to company logo image
            
        Returns:
            Path: Full path to the generated Excel report
        """
        # Step 1: Get employee/auditor information
        employee_data = SecurityAudit.get_form_data(window)
        
        # Step 2: Import and call all data collectors
        # (Re-importing here ensures fresh data and avoids circular imports)
        from collectors.imports import (
            get_antivirus_info, get_system_events, get_firewall_status,
            get_installed_applications, get_network_connections,
            get_powershell_history, get_running_processes, get_scheduled_tasks,
            get_startup_programs, get_system_info, get_usb_history, get_local_users
        )
        
        # Collect all security-related system information
        antivirus_info = get_antivirus_info()
        event_logs_info = get_system_events()
        firewall_status = get_firewall_status()
        installed_apps = get_installed_applications()
        network_connections = get_network_connections()
        powershell_history = get_powershell_history()
        running_process = get_running_processes()
        scheduled_task = get_scheduled_tasks()
        start_up_task = get_startup_programs()
        system_info = get_system_info()
        usb_history = get_usb_history()
        local_users = get_local_users()
        
        # Ensure output directory exists
        Path(output_folder).mkdir(exist_ok=True)
        
        # Generate unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Security_Audit_{employee_data['first_name']}_{employee_data['last_name']}_{timestamp}.xlsx"
        filepath = Path(output_folder) / filename

        # === Step 3: Dynamic Risk Level Assessment ===
        # Count potentially suspicious running processes
        suspicious_count = sum(1 for p in running_process 
                             if any(x in str(p).lower() for x in 
                                   ['powershell', 'cmd.exe', 'mshta', 'wscript', 'cscript']))
        
        firewall_str = str(firewall_status).upper()
        
        # Determine overall risk level based on multiple factors
        if suspicious_count >= 5 or "OFF" in firewall_str or "DISABLED" in firewall_str:
            risk_level = "CRITICAL"
        elif suspicious_count >= 3 or len(running_process) > 250:
            risk_level = "HIGH"
        elif suspicious_count >= 1:
            risk_level = "MEDIUM"
        else:
            risk_level = "LOW"

        risk_colors = {
            "LOW": "90EE90",      # Light green
            "MEDIUM": "FFB366",   # Orange
            "HIGH": "FF8C00",     # Dark orange
            "CRITICAL": "FF6666"  # Red
        }

        # === Step 4: Create Excel File with pandas ===
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            
            # Dashboard Sheet - High-level summary
            dashboard_data = {
                "Metric": ["Audit Date", "Auditor", "Position", "Department", "Computer Name",
                          "Total Installed Apps", "Running Processes", 
                          "Active Network Connections", "Risk Level"],
                "Value": [
                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                    f"{employee_data['first_name']} {employee_data['last_name']}",
                    employee_data.get('position', 'N/A'),
                    employee_data.get('Department', 'N/A'),   # Note: key matches form data
                    system_info.get('Computer Name', 'N/A'),
                    len(installed_apps),
                    len(running_process),
                    len(network_connections),
                    risk_level
                ]
            }
            pd.DataFrame(dashboard_data).to_excel(writer, sheet_name="Dashboard", index=False)

            # Employee Info Sheet
            pd.DataFrame([employee_data]).to_excel(writer, sheet_name="Employee Info", index=False)
            
            # Antivirus Information
            pd.DataFrame([antivirus_info.get("Windows_Defender", {})]).to_excel(
                writer, sheet_name="Windows Defender", index=False)
            
            if antivirus_info.get("Other_Antivirus"):
                pd.DataFrame(antivirus_info["Other_Antivirus"]).to_excel(
                    writer, sheet_name="Other Antivirus", index=False)
            else:
                pd.DataFrame([{"Message": "No other antivirus detected"}]).to_excel(
                    writer, sheet_name="Other Antivirus", index=False)
            
            # All other detailed sheets
            pd.DataFrame(event_logs_info).to_excel(writer, sheet_name="Windows Events", index=False)
            pd.DataFrame([{"Firewall Status": firewall_status}]).to_excel(writer, sheet_name="Firewall Status", index=False)
            pd.DataFrame(installed_apps).to_excel(writer, sheet_name="Installed Apps List", index=False)
            pd.DataFrame(network_connections).to_excel(writer, sheet_name="Network Connections", index=False)
            pd.DataFrame(powershell_history).to_excel(writer, sheet_name="Powershell History", index=False)
            pd.DataFrame(running_process).to_excel(writer, sheet_name="Running Processes", index=False)
            pd.DataFrame(scheduled_task).to_excel(writer, sheet_name="Scheduled Tasks", index=False)
            pd.DataFrame(start_up_task).to_excel(writer, sheet_name="Startup Programs", index=False)
            
            pd.DataFrame(list(system_info.items()), columns=["System Attribute", "Details"]).to_excel(
                writer, sheet_name="System Information", index=False)
            
            pd.DataFrame(usb_history).to_excel(writer, sheet_name="USB History", index=False)
            pd.DataFrame(local_users).to_excel(writer, sheet_name="Local Users", index=False)

        # === Step 5: Advanced Formatting & Charting with openpyxl ===
        wb = load_workbook(filepath)
        dashboard = wb["Dashboard"]

        # Highlight Risk Level cell with color coding
        risk_cell = dashboard["B10"]   # Risk Level value cell
        risk_color = risk_colors.get(risk_level, "90EE90")
        risk_cell.fill = PatternFill(start_color=risk_color, end_color=risk_color, fill_type="solid")
        risk_cell.font = Font(bold=True, color="000000")

        # === Prepare Top 10 Running Processes Data ===
        process_list = []
        for p in running_process:
            if isinstance(p, dict):
                name = p.get('Name') or p.get('name', 'Unknown')
            else:
                name = str(p)
            if name and name != 'None':
                clean_name = str(name).split('.')[0][:25]  # Clean and truncate
                process_list.append(clean_name)

        # Count occurrences and get top 10
        top_10 = Counter(process_list).most_common(10)

        # Write Top 10 data to Dashboard (columns G and H)
        dashboard['G1'] = "Process Name"
        dashboard['H1'] = "Count"
        
        for i, (proc, count) in enumerate(top_10, start=2):
            dashboard.cell(row=i, column=7, value=proc)
            dashboard.cell(row=i, column=8, value=count)

        # === Create Bar Chart for Top 10 Processes ===
        chart = BarChart()
        chart.title = "Top 10 Running Processes"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Process Name"
        chart.style = 2

        # Define data ranges for the chart
        data = Reference(dashboard, min_col=8, min_row=1, max_row=len(top_10) + 1)
        categories = Reference(dashboard, min_col=7, min_row=2, max_row=len(top_10) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Chart appearance settings
        chart.height = 15
        chart.width = 26
        chart.legend = None
        chart.x_axis.tickLblPos = "low"   # Better label positioning

        # Add chart to Dashboard sheet
        dashboard.add_chart(chart, "J2")

        # Apply consistent styling to ALL sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            SecurityAudit._style_worksheet(ws, is_dashboard=(sheet_name == "Dashboard"))

        # Add company logo to Dashboard (if provided)
        if logo_path and Path(logo_path).exists():
            try:
                img = Image(logo_path)
                img.width = 180
                img.height = 90
                dashboard.add_image(img, 'A1')
            except Exception as e:
                print(f"⚠️ Could not add logo: {e}")

        # Save the final formatted workbook
        wb.save(filepath)
        
        
        
        return filepath
    
    #Author: James Ian Largo
    #Purpose: For Future Use